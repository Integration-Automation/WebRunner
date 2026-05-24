"""
下載檔案驗證:PDF/CSV/Excel/JSON/檔名/SHA256 比對,給 E2E 測試用。

Download verification helpers. Pairs with the existing browser download
action commands: after a test triggers a file save, these utilities
poll the download directory, extract content (PDF text / CSV rows /
Excel rows / JSON), and assert on it.

Soft dependencies — only required when the matching extractor is used:

* PDF text → ``pypdf`` (preferred) or ``pdfplumber``
* Excel rows → ``openpyxl``

CSV and JSON use the standard library.
"""
from __future__ import annotations

import csv
import hashlib
import json
import re
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Optional, Pattern, Union

from je_web_runner.utils.exception.exceptions import WebRunnerException
from je_web_runner.utils.logging.loggin_instance import web_runner_logger


class DownloadVerifyError(WebRunnerException):
    """Raised on missing files, parse errors, or assertion failures."""


# ---------- waiting -------------------------------------------------------

def wait_for_download(
    download_dir: Union[str, Path],
    *,
    pattern: Union[str, Pattern[str]] = r".+",
    timeout: float = 30.0,
    poll_interval: float = 0.5,
    stable_for: float = 0.5,
    sleep_fn: Callable[[float], None] = time.sleep,
    time_fn: Callable[[], float] = time.time,
    exclude_extensions: Iterable[str] = (".crdownload", ".part", ".tmp"),
) -> Path:
    """
    等到 download_dir 內出現符合 pattern 的「完成」檔案。
    Poll a download dir until a file matches ``pattern`` AND has stayed
    the same size for ``stable_for`` seconds. Skips partial-download
    suffixes (``.crdownload`` etc.).

    Returns the absolute path to the file.
    """
    directory = Path(download_dir)
    if not directory.is_dir():
        raise DownloadVerifyError(f"download dir not found: {directory}")
    regex = re.compile(pattern) if isinstance(pattern, str) else pattern
    excluded = tuple(e.lower() for e in exclude_extensions)
    start = time_fn()
    last_seen_size: Dict[str, int] = {}
    last_seen_time: Dict[str, float] = {}
    while True:
        for entry in directory.iterdir():
            if not entry.is_file():
                continue
            name = entry.name
            if any(name.lower().endswith(suffix) for suffix in excluded):
                continue
            if not regex.search(name):
                continue
            size = entry.stat().st_size
            prev_size = last_seen_size.get(name)
            now = time_fn()
            if prev_size == size:
                if now - last_seen_time.get(name, now) >= stable_for and size > 0:
                    web_runner_logger.info(f"wait_for_download: matched {entry}")
                    return entry.resolve()
            else:
                last_seen_size[name] = size
                last_seen_time[name] = now
        if time_fn() - start >= timeout:
            raise DownloadVerifyError(
                f"timeout waiting for {pattern!r} in {directory} after {timeout}s"
            )
        sleep_fn(poll_interval)


# ---------- hashing -------------------------------------------------------

def sha256_of_file(path: Union[str, Path], *, chunk_size: int = 65_536) -> str:
    """Stream-hash a file with SHA-256."""
    p = Path(path)
    if not p.is_file():
        raise DownloadVerifyError(f"not a file: {p}")
    hasher = hashlib.sha256()
    with open(p, "rb") as fp:
        while True:
            chunk = fp.read(chunk_size)
            if not chunk:
                break
            hasher.update(chunk)
    return hasher.hexdigest()


def assert_file_sha256(path: Union[str, Path], expected: str) -> None:
    """Raise unless ``path``'s SHA-256 equals ``expected`` (case-insensitive)."""
    actual = sha256_of_file(path)
    if actual.lower() != expected.lower():
        raise DownloadVerifyError(
            f"sha256 mismatch for {path}: expected {expected}, got {actual}"
        )


# ---------- PDF -----------------------------------------------------------

def extract_pdf_text(
    path: Union[str, Path],
    *,
    page_separator: str = "\n",
) -> str:
    """
    用 pypdf / pdfplumber 抽出整份 PDF 的文字。
    Concatenate per-page text. ``pypdf`` is tried first (lighter, pure
    Python); ``pdfplumber`` is the fallback. Raises if neither is
    installed or the file isn't a valid PDF.
    """
    p = Path(path)
    if not p.is_file():
        raise DownloadVerifyError(f"PDF not found: {p}")
    try:
        from pypdf import PdfReader  # type: ignore[import-not-found]
        reader = PdfReader(str(p))
        return page_separator.join(
            (page.extract_text() or "") for page in reader.pages
        )
    except ImportError:
        pass
    try:
        import pdfplumber  # type: ignore[import-not-found]
        pieces: List[str] = []
        with pdfplumber.open(str(p)) as pdf:
            for page in pdf.pages:
                pieces.append(page.extract_text() or "")
        return page_separator.join(pieces)
    except ImportError as error:
        raise DownloadVerifyError(
            "PDF text extraction requires pypdf or pdfplumber. "
            "Install one: pip install pypdf"
        ) from error
    except Exception as error:  # noqa: BLE001 — library-specific parse errors
        raise DownloadVerifyError(f"failed to extract PDF text from {p}: {error!r}") from error


def assert_pdf_contains(path: Union[str, Path], substring: str) -> None:
    """Raise if the extracted PDF text doesn't include ``substring``."""
    text = extract_pdf_text(path)
    if substring not in text:
        raise DownloadVerifyError(
            f"PDF {path} does not contain substring {substring!r}"
        )


def assert_pdf_matches(path: Union[str, Path], pattern: Union[str, Pattern[str]]) -> str:
    """Raise unless the PDF text matches ``pattern``; returns the match."""
    text = extract_pdf_text(path)
    regex = re.compile(pattern) if isinstance(pattern, str) else pattern
    match = regex.search(text)
    if match is None:
        raise DownloadVerifyError(
            f"PDF {path} does not match pattern {pattern!r}"
        )
    return match.group(0)


# ---------- CSV -----------------------------------------------------------

def read_csv_rows(
    path: Union[str, Path],
    *,
    encoding: str = "utf-8-sig",
    dialect: str = "excel",
) -> List[Dict[str, str]]:
    """Read a CSV file as a list of dicts (header-driven)."""
    p = Path(path)
    if not p.is_file():
        raise DownloadVerifyError(f"CSV not found: {p}")
    try:
        with open(p, encoding=encoding, newline="") as fp:
            reader = csv.DictReader(fp, dialect=dialect)
            return [dict(row) for row in reader]
    except (OSError, csv.Error) as error:
        raise DownloadVerifyError(f"cannot read CSV {p}: {error!r}") from error


def assert_csv_columns(path: Union[str, Path], expected_columns: Iterable[str]) -> None:
    """Raise if the CSV is missing any of ``expected_columns``."""
    rows = read_csv_rows(path)
    if not rows:
        raise DownloadVerifyError(f"CSV {path} is empty")
    present = set(rows[0].keys())
    missing = [c for c in expected_columns if c not in present]
    if missing:
        raise DownloadVerifyError(
            f"CSV {path} missing columns {missing} (have {sorted(present)})"
        )


def assert_csv_row_count(
    path: Union[str, Path],
    *,
    minimum: Optional[int] = None,
    maximum: Optional[int] = None,
) -> int:
    """Raise unless the row count is within bounds. Returns the actual count."""
    count = len(read_csv_rows(path))
    if minimum is not None and count < minimum:
        raise DownloadVerifyError(
            f"CSV {path} has {count} rows, expected >= {minimum}"
        )
    if maximum is not None and count > maximum:
        raise DownloadVerifyError(
            f"CSV {path} has {count} rows, expected <= {maximum}"
        )
    return count


# ---------- Excel ---------------------------------------------------------

def read_excel_rows(
    path: Union[str, Path],
    *,
    sheet: Optional[Union[str, int]] = None,
) -> List[Dict[str, Any]]:
    """
    讀 .xlsx 為 list of dict (假設第一列是 header)。
    Read an Excel sheet (defaults to the first/active one). Requires
    ``openpyxl``; raises with an install hint when missing.
    """
    p = Path(path)
    if not p.is_file():
        raise DownloadVerifyError(f"Excel file not found: {p}")
    try:
        from openpyxl import load_workbook  # type: ignore[import-not-found]
    except ImportError as error:
        raise DownloadVerifyError(
            "Excel extraction requires openpyxl. Install: pip install openpyxl"
        ) from error
    try:
        wb = load_workbook(filename=str(p), read_only=True, data_only=True)
    except Exception as error:  # noqa: BLE001 — openpyxl raises many types
        raise DownloadVerifyError(f"cannot open {p}: {error!r}") from error
    try:
        if sheet is None:
            ws = wb.active
        elif isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header)]
        return [
            {headers[i]: value for i, value in enumerate(row) if i < len(headers)}
            for row in rows_iter
        ]
    finally:
        wb.close()


# ---------- JSON ----------------------------------------------------------

def read_json_file(path: Union[str, Path]) -> Any:
    p = Path(path)
    if not p.is_file():
        raise DownloadVerifyError(f"JSON file not found: {p}")
    try:
        with open(p, encoding="utf-8") as fp:
            return json.load(fp)
    except (OSError, ValueError) as error:
        raise DownloadVerifyError(f"cannot read JSON {p}: {error!r}") from error


def assert_json_matches_schema(path: Union[str, Path], schema: Dict[str, Any]) -> None:
    """
    用簡化版 schema 驗證 JSON:``{"key": "type" or {nested}}``。
    Minimal schema validator (NO jsonschema dependency). Schema example::

        {"name": "str", "age": "int", "address": {"city": "str"}}

    Raises on mismatch; type names are Python type aliases: ``str``,
    ``int``, ``float``, ``bool``, ``list``, ``dict``.
    """
    payload = read_json_file(path)
    _check_schema(payload, schema, path=str(path))


_TYPE_ALIASES: Dict[str, type] = {
    "str": str,
    "int": int,
    "float": float,
    "bool": bool,
    "list": list,
    "dict": dict,
    "any": object,
}


def _check_schema(payload: Any, schema: Any, *, path: str, prefix: str = "$") -> None:
    if isinstance(schema, str):
        expected = _TYPE_ALIASES.get(schema)
        if expected is None:
            raise DownloadVerifyError(f"unknown schema type {schema!r} at {prefix}")
        if expected is object:
            return
        if not isinstance(payload, expected):
            raise DownloadVerifyError(
                f"JSON {path} at {prefix}: expected {schema}, got {type(payload).__name__}"
            )
        return
    if isinstance(schema, dict):
        if not isinstance(payload, dict):
            raise DownloadVerifyError(
                f"JSON {path} at {prefix}: expected object, got {type(payload).__name__}"
            )
        for key, sub_schema in schema.items():
            if key not in payload:
                raise DownloadVerifyError(
                    f"JSON {path} at {prefix}: missing key {key!r}"
                )
            _check_schema(payload[key], sub_schema, path=path, prefix=f"{prefix}.{key}")
        return
    raise DownloadVerifyError(f"unsupported schema node at {prefix}: {schema!r}")


# ---------- one-shot all-in-one ------------------------------------------

@dataclass
class DownloadAssertion:
    """All the constraints :func:`assert_download` can check at once."""

    filename_pattern: Optional[Union[str, Pattern[str]]] = None
    sha256: Optional[str] = None
    pdf_contains: Optional[str] = None
    csv_columns: Optional[List[str]] = None
    json_schema: Optional[Dict[str, Any]] = None
    min_size_bytes: Optional[int] = None
    max_size_bytes: Optional[int] = None


def assert_download(path: Union[str, Path], assertion: DownloadAssertion) -> None:
    """
    一次跑完整套 download 驗證,任何一條不過就 raise。
    Combined check that walks every populated field of ``assertion`` and
    raises on the first failure. Use this when a download has multiple
    constraints (filename + content + hash).
    """
    p = Path(path)
    if not p.is_file():
        raise DownloadVerifyError(f"download file missing: {p}")
    if assertion.filename_pattern is not None:
        regex = (
            re.compile(assertion.filename_pattern)
            if isinstance(assertion.filename_pattern, str)
            else assertion.filename_pattern
        )
        if regex.search(p.name) is None:
            raise DownloadVerifyError(
                f"filename {p.name!r} does not match {assertion.filename_pattern!r}"
            )
    size = p.stat().st_size
    if assertion.min_size_bytes is not None and size < assertion.min_size_bytes:
        raise DownloadVerifyError(
            f"file {p} size {size} < min {assertion.min_size_bytes}"
        )
    if assertion.max_size_bytes is not None and size > assertion.max_size_bytes:
        raise DownloadVerifyError(
            f"file {p} size {size} > max {assertion.max_size_bytes}"
        )
    if assertion.sha256 is not None:
        assert_file_sha256(p, assertion.sha256)
    if assertion.pdf_contains is not None:
        assert_pdf_contains(p, assertion.pdf_contains)
    if assertion.csv_columns is not None:
        assert_csv_columns(p, assertion.csv_columns)
    if assertion.json_schema is not None:
        assert_json_matches_schema(p, assertion.json_schema)
